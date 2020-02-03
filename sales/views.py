from django.shortcuts import render, redirect,get_object_or_404
from .models import *
from sales.forms import *
from products.models import *
from sales.models import Sale

from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def sales(request):
    sales = Sale.objects.all()
    return render(request, 'sales/all_sales.html', locals())


def sales_report(request):
    sales = Sale.objects.all()
    total = [i.product.product_price * i.quantity for i in sales]
    final_total = sum(total)
    return render(request, 'reports/sales_report.html', locals())

def add_item(request,cls):
    if request.method == "POST":
        form = cls(request.POST)
        if form.is_valid():
            obj = Product.objects.filter(id=request.POST['product']).first()
            obj.product_qyt = obj.product_qyt - int(request.POST['quantity'])
            obj.save()
            form.save()
            return redirect('sales')
    else:
        form = cls()
        return render(request, 'sales/add_sell.html', locals())


def add_sell(request):
    return add_item(request, SalesForm)

def export_salesreport_to_xlsx(request):
    """
    Downloads all purchases as Excel file with a single worksheet
    """
    salesreport_queryset = Sale.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-salesreport.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

   
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Sales Report'

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the titles for columns
    columns = [
        ('Sales Date', 20),
        ('Product Name', 15),
        ('Client Name', 15),
        ('Product Category', 10),
        ('Amount(KSH)', 15),
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all purchases
    for sale in salesreport_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            sale.date,
            sale.product.product_name,
            sale.client.client_name,
            sale.product.product_category,
            sale.get_total(),
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value  in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return (response)

def filter(request):
    sales = Sale.objects.filter(date__range=(
        request.POST['start_date'],
        request.POST['end_date']
    ))
    
    total = [i.product.product_price * i.quantity for i in sales]
    final_total = sum(total)
    
    return render(request, 'reports/sales_report.html',locals())




