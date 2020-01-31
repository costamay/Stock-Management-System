from django.shortcuts import render, redirect,get_object_or_404
import datetime as dt
from .models import *
from .forms import *
from sales.models import *

from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def all_suppliers(request):
    suppliers = Supplier.objects.all()
    total_suppliers = Supplier.objects.all().count()
    return render(request, 'supplier/manage_supplier.html', locals())

def add_item(request, cls):
        
    if request.method == "POST":
        form = cls(request.POST)
        if form.is_valid():
            form.save()
            return redirect('all_suppliers')
    else:
        form = cls()
        return render(request, 'add_new_supplier.html', locals())

def add_supplier(request):
    return add_item(request, SupplierForm)

def edit_item(request, pk, model, cls):
    item = get_object_or_404(model, pk=pk)

    if request.method == "POST":
        form = cls(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('all_suppliers')
    else:
        form = cls(instance=item)

        return render(request, 'edit_supplier.html', locals())

def edit_supplier(request, pk):
    return edit_item(request, pk, Supplier, SupplierForm)

def delete_supplier(request, pk):

    template = 'supplier/manage_supplier.html'
    Supplier.objects.filter(id=pk).delete()
    suppliers = Supplier.objects.all()
    
    return render(request, template, locals())


def purchase_report(request):
    purchases = Supplier.objects.all()
    total = [i.materials.price * i.materials.quantity for i in purchases]
    final_total = sum(total)
   
    return render(request, 'reports/purchase_report.html', locals())

def todays_purchase(request):
        date = dt.date.today()
        purchases = Supplier.todays_purchase()
        total_purchase = [i.materials.price * i.materials.quantity for i in purchases]
        final_total_purchase = sum(total_purchase)

        sales = Sale.todays_sales()
        total_sales = [i.product.product_price * i.quantity for i in sales]
        final_total_sales = sum(total_sales)

        return render(request, 'reports/todays_report.html', locals())

def filter_purchase(request):
    purchases = Supplier.objects.filter(date__range=(
        request.POST['start_date'],
        request.POST['end_date']
    ))
   
    total = [i.materials.price * i.materials.quantity for i in purchases]
    final_total = sum(total)
    return render(request, 'reports/purchase_report.html',locals())

def search_suppliers(request):

    if 'search' in request.GET and request.GET["search"]:
        search_term = request.GET.get("search")
        searched_suppliers = Supplier.search(search_term)
        message = f"{search_term}"
        return render(request, 'supplier/search.html',{"message":message,"suppliers": searched_suppliers})
    else:
        message = "You haven't searched for any term"
        return render(request, 'supplier/search.html',{"message":message})

def export_purchasesreport_to_xlsx(request):
    """
    Downloads all purchases as Excel file with a single worksheet
    """
    purchasesreport_queryset = Supplier.objects.all()
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-purchasesreport.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
#     workbook.remove(workbook.active)
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Purchases Report'

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the titles for columns
    columns = [
        ('Purchase Date', 20),
        ('Purchase No', 10),
        ('Supplier Name', 15),
        ('Amount(KSH)', 15),
        
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, (column_title, column_width) in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        # cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = column_width

    # Iterate through all purchases
    for purchase in purchasesreport_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            purchase.date,
            purchase.id,
            purchase.supplier_name,
            purchase.get_total(),   
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = wrapped_alignment
    workbook.save(response)
    return response